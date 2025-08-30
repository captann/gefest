from openpyxl import load_workbook
import requests
from sqlalchemy import create_engine
from typing import List, Tuple, Optional, Dict
from models import *
from split_location import parse
from sqlalchemy.orm import sessionmaker
import os
from urllib.parse import quote_plus

from app_config import remote, creds_path, YANDEX_MAPS_API_KEY, DB_FILE

if remote:
    pass
else:
    sqlite_engine = create_engine(f'sqlite:///{DB_FILE}')

    Session = sessionmaker(bind=sqlite_engine)




def unsuccessful(message):
    return {"success": False,
            "message": message}

def successful(message):
    return {"success": True,
            "message": message}



"""DATABASE_URL = f"sqlite:///{DB_FILE}"  # ← путь к твоему .db файлу
engine = create_engine(DATABASE_URL, echo=False)
Session = sessionmaker(bind=engine)"""

def geocode_address(api_key: str, address: str) -> Dict:
    """
    Геокодирует адрес с помощью API Яндекс.Карт

    :param api_key: ваш API ключ Яндекс.Карт
    :param address: адрес для геокодирования
    :return: кортеж (долгота, широта)
    """
    url = "https://geocode-maps.yandex.ru/1.x/"
    params = {
        "apikey": api_key,
        "geocode": address,
        "format": "json"
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()

        data = response.json()
    except Exception as e:
        return {"success": False,
                "message": f"Не удалось выполнить запрос к геокодеру. Ошибка: {str(e.args)}"}
    try:
        pos = data['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']
        lon, lat = map(float, pos.split())
        return {"success": True,
                "lon": lon,
                "lat": lat}
    except Exception as e:

        return {"success": False,
                "message": f"Не удалось геокодировать адрес: {address}\nОшибка: {str(e.args)}"}


def is_point_in_area(point: Tuple[float, float], polygon_ids: List) -> bool:
    """
    Проверяет, находится ли точка внутри полигона с помощью алгоритма "ray casting"

    :param point: кортеж (долгота, широта) проверяемой точки
    :param polygon: список кортежей (долгота, широта) вершин полигона
    :return: True если точка внутри полигона, иначе False
    """
    if not polygon_ids:
        return True
    poligons_points = []
    with Session() as session:
        for pol_id in polygon_ids:
            record = session.query(PolygonModel).filter_by(polygon_id=pol_id).first()
            if record:
                lon = [float(x) for x in record.lon_points.split(';')]
                lat = [float(x) for x in record.lat_points.split(';')]
                poligons_points.append([[lon[i], lat[i]] for i in range(len(lat))])
    x, y = point
    res = []
    for polygon in poligons_points:
        n = len(polygon)
        inside = False
        p1x, p1y = polygon[0]
        for i in range(n + 1):
            p2x, p2y = polygon[i % n]
            if y > min(p1y, p2y):
                if y <= max(p1y, p2y):
                    if x <= max(p1x, p2x):
                        if p1y != p2y:
                            xinters = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                        if p1x == p2x or x <= xinters:
                            inside = not inside
            p1x, p1y = p2x, p2y
        res.append(inside)
    return any(res)

class User:
    def __init__(self, login, password):

        self.login = login
        self.password = password
        self.get_id()

    def get_polygons(self):
        """Возвращает список объектов Polygon, принадлежащих пользователю"""
        r = []
        with Session() as session:
            records = session.query(PolygonModel).filter_by(user_id=self.user_id, deleted=0).all()
            polygon_ids = [p.polygon_id for p in records]

        for pid in polygon_ids:
            polygon = Polygon.get_by_id(pid)
            if polygon:
                r.append(polygon)
        return r

    def get_checked_polygons_ids(self):
        """Возвращает список ID отмеченных полигонов пользователя"""
        r = []
        with Session() as session:
            records = session.query(PolygonModel).filter_by(
                user_id=self.user_id,
                deleted=0,
                checked=1
            ).all()
            r = [p.polygon_id for p in records]
        return r

    @classmethod
    def get_by_id(cls, user_id):
        """Возвращает объект User по ID, если он существует"""
        user_obj = None
        with Session() as session:
            record = session.query(UserModel).filter_by(user_id=user_id).first()
            if record:
                user_obj = cls(record.login, record.password)
        return user_obj

    def get_polygons_ids(self):
        """Возвращает список ID всех полигонов пользователя (не удалённых)"""
        r = []
        with Session() as session:
            records = session.query(PolygonModel).filter_by(
                user_id=self.user_id,
                deleted=0
            ).all()
            r = [record.polygon_id for record in records]
        return r

    def get_id(self):
        """Получает user_id по логину, либо создаёт нового пользователя"""
        with Session() as session:
            try:
                user = session.query(UserModel).filter_by(login=self.login).first()

                if not user:
                    new_user = UserModel(
                        login=self.login,
                        password=self.password,
                        role='guest'
                    )
                    session.add(new_user)
                    session.commit()
                    session.refresh(new_user)
                    self.user_id = new_user.user_id
                    self.role = new_user.role
                else:
                    self.user_id = user.user_id
                    self.role = user.role
            except Exception as e:
                session.rollback()
                print(f"Ошибка при получении или создании пользователя: {str(e)}")

    def __dict__(self):
        r = {}
        r['user_id'] = self.user_id
        r['login'] = self.login
        r['role'] = self.role
        return r

class Polygon:
    def __init__(self, user_id, polygon_name, points, color, checked):
        self.user_id = user_id
        self.polygon_name = polygon_name
        self.color = color
        self.points = points
        self.lon_points = ';'.join([str(i[0]) for i in self.points])
        self.lat_points = ';'.join([str(i[1]) for i in self.points])
        self.polygon_id = None
        self.checked = checked
        self.get_id()

    @staticmethod
    def delete(polygon_id):
        """Помечает полигон как удалённый (deleted = 1)"""
        with Session() as session:
            try:
                polygon = session.query(PolygonModel).filter_by(polygon_id=polygon_id).first()
                if polygon:
                    polygon.deleted = 1
                    session.commit()
            except Exception as e:
                session.rollback()
                print(f"Ошибка при удалении полигона {polygon_id}: {str(e)}")

    def get_id(self):
        """Получает ID полигона по имени и user_id. Если не найден, создаёт новый."""
        with Session() as session:
            try:
                polygon = session.query(PolygonModel).filter_by(
                    polygon_name=self.polygon_name,
                    user_id=self.user_id,
                    deleted=0
                ).first()

                if not polygon:
                    new_polygon = PolygonModel(
                        user_id=self.user_id,
                        polygon_name=self.polygon_name,
                        color=self.color,
                        lon_points=self.lon_points,
                        lat_points=self.lat_points,
                        deleted=0,
                        checked=self.checked
                    )
                    session.add(new_polygon)
                    session.commit()
                    session.refresh(new_polygon)  # обновляем, чтобы получить polygon_id
                    self.polygon_id = new_polygon.polygon_id
                else:
                    self.polygon_id = polygon.polygon_id
            except Exception as e:
                session.rollback()
                print(f"Ошибка при получении или создании полигона: {str(e)}")

    @staticmethod
    def check_color(color, user_id):
        """Проверяет, существует ли полигон с таким цветом для данного пользователя."""
        result = False
        with Session() as session:
            try:
                count = session.query(PolygonModel).filter_by(
                    color=color,
                    user_id=user_id,
                    deleted=0
                ).count()
                result = count > 0
            except Exception as e:
                print(f"Ошибка при проверке цвета: {str(e)}")
        return result

    @staticmethod
    def check_name(name, user_id):
        """Проверяет, существует ли полигон с таким именем для данного пользователя."""
        result = False
        with Session() as session:
            try:
                count = session.query(PolygonModel).filter_by(
                    polygon_name=name,
                    user_id=user_id,
                    deleted=0
                ).count()
                result = count > 0
            except Exception as e:
                print(f"Ошибка при проверке имени полигона: {str(e)}")
        return result

    @classmethod
    def get_by_id(cls, polygon_id):
        """Возвращает объект Polygon по его ID, если он существует."""
        polygon_obj = None
        with Session() as session:
            record = session.query(PolygonModel).filter_by(polygon_id=polygon_id).first()

            if record:

                lon_list = [float(x) for x in record.lon_points.split(';')]
                lat_list = [float(x) for x in record.lat_points.split(';')]
                points = list(zip(lon_list, lat_list))
                polygon_obj = cls(
                    record.user_id,
                    record.polygon_name,
                    points,
                    record.color,
                    record.checked
                )
        return polygon_obj

    def __str__(self):
        """Возвращает строковое представление полигона."""
        return f"{self.polygon_name}\n" + "\n".join([' '.join(str(i) for i in point) for point in self.points])

class Address:
    def __init__(self, **kwargs):
        self.is_valid = False
        self.submit_required = False
        self.message = ""
        if kwargs.get("raw_address", False):
            result = parse(kwargs.get("raw_address", ""))
            if not result["success"]:
                self.message = result["message"] + f"; {kwargs.get('raw_address', '')}"
            else:
                self.home_id = result["home_id"]
                self.home_name = result["home_name"]
                self.home_address = result["home_address"]
                self.print = kwargs.get("print", 0)
                self.sign = kwargs.get("sign", 0)
                is_exist = self.is_exist(self.home_id)

                if not is_exist:
                    geocoding = geocode_address(YANDEX_MAPS_API_KEY, self.home_address)
                    self.submit_required = True
                    if not geocoding["success"]:
                        self.message = geocoding["message"] + f"; {self.home_address}"
                    else:
                        self.lon = geocoding["lon"]
                        self.lat = geocoding["lat"]
                        attempt = {"success": True,
                                   "message": "ok"}
                        if not attempt.get("success", False):
                            self.message = attempt["message"]

                        self.is_valid = attempt.get("success", False)
                else:
                    info = self.get_by_id(self.home_id)
                    self.lon = info["lon"]
                    self.lat = info["lat"]
                    self.is_valid = info["is_valid"]
                    self.stuff_id = info["stuff_id"]
                    self.print = info["print"]
                    self.sign = info["sign"]
        elif kwargs.get("home_id", False):
            self.is_valid = Address.is_exist(kwargs.get("home_id", False))
            if not self.is_valid:
                self.message = f"здания с ID {kwargs.get('home_id', -1)} нет"
            else:
                r = self.get_by_id(kwargs.get("home_id", -1))
                self.is_valid = r["is_valid"]
                self.home_id = r["home_id"]
                self.stuff_id = r["stuff_id"]
                self.home_name = r["home_name"]
                self.home_address = r["home_address"]
                self.lon = r["lon"]
                self.lat = r["lat"]
                self.print = r["print"]
                self.sign = r["sign"]


    @staticmethod
    def is_exist(home_id):
        with Session() as session:
            r = session.query(AddressModel).filter_by(home_id=home_id).first() is not None
        return r

    @staticmethod
    def get_by_id(home_id):
        with Session() as session:
            record = session.query(AddressModel).filter_by(home_id=home_id).first()
        if record:
            # Возвращаем простой объект-обёртку или словарь
            return {
                "is_valid": True,
                "home_id": record.home_id,
                "home_name": record.home_name,
                "home_address": record.home_address,
                "lon": record.lon,
                "lat": record.lat,
                "stuff_id": record.stuff_id,
                "print": record.print,
                "sign": record.sign
            }
        else:
            return {"is_valid": False}

    def make_contribution(self):
        try:
            with Session() as session:
                new_adress = AddressModel(
                    home_id = self.home_id,
                    home_name=self.home_name,
                    home_address=self.home_address,
                    lon=self.lon,
                    lat=self.lat,
                    print=self.print,
                    sign=self.sign
                )
                session.add(new_adress)
                session.commit()
            return {"success": True}
        except Exception as e:
            return {
                unsuccessful(f"Ошибка добавления адреса: ошибка {str(e.args)}")
            }

    def __str__(self):
        if not self.is_valid:
            return f"Invalid\n{self.message}"
        return f"home_id: {self.home_id}\nhome_name:{self.home_name}\nhome_address:{self.home_address}\nlon:{self.lon}\nlat:{self.lat}"

    def __dict__(self):
        if not self.is_valid:
            return {"valid": False, "message": self.message}
        return {"home_id": self.home_id,
                "home_name":self.home_name,
                "home_address":self.home_address,
                "lon":self.lon,
                "lat":self.lat,
                "print":self.print,
                "sign":self.sign}

class InvalidAddressError(Exception):
    pass

class Task:
    def __init__(self, task_id, date, address: Address, problem, solution, blank=False, archieved=False, ppr=False):
        if not address.is_valid:
            raise InvalidAddressError("Invalid address")
        self.archieved = archieved
        self.task_id = task_id
        self.blank = blank
        self.address = address
        self.correct_date(date)
        self.correct_location(address)
        self.get_lon_lan()
        self.problem = problem
        self.solution = solution
        self.ppr = ppr
        self._save_task_to_db()


    def correct_date(self, date):
        r = ' '.join(date.split(' ')[1:-1])
        if r:
            self.date = r
        else:
            self.date = date



    def correct_location(self, address: Address):
        self.R = {
                "address": address.home_address,
                "name": address.home_name,
                "home_id": address.home_id
            }


        self.location = self.R['address']
        self.origin = str(self.R['home_id']) + " " + self.R['name'] + " " + self.R['address']
        self.home_id = self.R['home_id']

    def _save_task_to_db(self):
        """Сохраняет задачу в базу данных, если она ещё не существует"""
        with Session() as session:
            try:
                existing_task = session.query(TaskModel).filter_by(task_id=self.task_id).first()
                self.already_exist = bool(existing_task)
                if self.already_exist:
                    record = session.query(TaskModel).filter(
                        TaskModel.task_id == self.task_id,
                        TaskModel.is_ppr == 0
                    ).first()
                    # проверка, что сущеуствущая задача - заявка, а не ппр
                    # если задача - заявка, но она запрашивается как ППР - до для этой ППР создаем новую
                    if record and self.ppr:
                        print(1)
                        self.already_exist = False
                        """record.task_id = generate_ppr_id(self.home_id)
                        session.commit()"""
                        self.task_id = generate_ppr_id(self.home_id)
                if not self.already_exist:
                    new_task = TaskModel(
                        task_id=self.task_id,
                        date=self.date,
                        home_id=self.home_id,
                        problem=self.problem,
                        solution=self.solution,
                        blank=self.blank,
                        archieved=self.archieved,  # ← добавлено
                        is_ppr=self.ppr
                    )
                    session.add(new_task)
                    session.commit()
                else:
                    self.archieved = existing_task.archieved
                    self.blank = existing_task.blank
                    self.ppr = existing_task.is_ppr
            except Exception as e:
                session.rollback()
                print(f"Ошибка при сохранении задачи {self.task_id}: {str(e)}")

    def get_lon_lan(self):
        """Получает координаты из БД или геокодера"""
        # Инициализируем БД при первом вызове
        # Пытаемся получить координаты из БД
        cached_coords = self._get_coords_from_db()
        if cached_coords:
            self.lon_lan = cached_coords
            return
        # Если в БД нет, используем геокодер
        coords = geocode_address(YANDEX_MAPS_API_KEY, self.location)
        self.lon_lan = (coords["lon"], coords["lat"])

        # Сохраняем в БД для будущих запросов

    def _get_coords_from_db(self) -> Optional[Tuple[float, float]]:
        """Пытается получить координаты из БД по home_id"""
        result = None
        with Session() as session:
            address = session.query(AddressModel).filter_by(home_id=self.R["home_id"]).first()
            if address:
                result = (address.lon, address.lat)
        return result

    def update_blank_status(self, is_ready: bool):
        """Обновляет статус акта о выполнении в базе данных"""
        with Session() as session:
            try:
                task = session.query(TaskModel).filter_by(task_id=self.task_id).first()
                if task:
                    task.blank = is_ready
                    session.commit()
                    self.is_green = is_ready
            except Exception as e:
                session.rollback()
                print(f"Ошибка при обновлении статуса задачи {self.task_id}: {str(e)}")



    @staticmethod
    def get_tasks(archieved=False, task_type=0):
        res = []
        with Session() as session:
            all_tasks = session.query(TaskModel).all()
            for t in all_tasks:
                address_row = session.query(AddressModel).filter_by(home_id=t.home_id).first()
                if not address_row:
                    continue  # адреса нет — пропускаем

                address = Address(
                    home_id=address_row.home_id,
                    home_name=address_row.home_name,
                    home_address=address_row.home_address,
                    lon=address_row.lon,
                    lat=address_row.lat
                )
                task = Task(
                    task_id=t.task_id,
                    date=t.date,
                    address=address,
                    problem=t.problem,
                    solution=t.solution,
                    blank=t.blank,
                    archieved=t.archieved,
                    ppr=t.is_ppr
                )

                if task_type == 0:
                    if archieved:
                        if not task.archieved:
                            res.append(task)
                    else:
                        res.append(task)
                elif task_type == 1 and not task.ppr:
                    if archieved:
                        if not task.archieved:
                            res.append(task)
                    else:
                        res.append(task)
                elif task_type == 2 and task.ppr:
                    if archieved:
                        if not task.archieved:
                            res.append(task)
                    else:
                        res.append(task)

        return res

    def __dict__(self):
        r = {"task_id": self.task_id,
             "date": self.date,
             "address": self.location,
             "home_id": self.home_id,
             "solution": self.solution,
             "problem": self.problem,
             "lan": self.lon_lan[1],
             "lon": self.lon_lan[0],
             "blank": int(self.blank),
             "archieved": int(self.archieved)
             }
        return r


def read_selected_columns(file_path: str, sheet_name: str = None,
                          start_row: int = 2, end_row: int = None,
                          columns_to_read: List[str] = ['A', 'B', "C", "D", "E", "F", "G", "H"]) -> List[Dict]:
    """
    Читает указанные столбцы из Excel-файла и возвращает список словарей

    :param file_path: путь к файлу Excel
    :param sheet_name: имя листа (если None, берётся первый лист)
    :param start_row: начальная строка данных (по умолчанию 2, пропускает шапку)
    :param end_row: конечная строка (если None, читает до последней заполненной строки)
    :param columns_to_read: список столбцов для чтения (по умолчанию A, C, D, F)
    :return: список словарей, где ключи - буквы столбцов, значения - данные ячеек
    """
    # Загружаем книгу
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)

    # Выбираем лист
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    # Определяем последнюю строку, если не задана
    if end_row is None:
        end_row = sheet.max_row

    result = []

    # Проходим по строкам
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        row_data = {}
        for col_idx, cell_value in enumerate(row, start=1):
            col_letter = chr(64 + col_idx)  # Преобразуем индекс в букву (1->A, 2->B и т.д.)
            if col_letter in columns_to_read:
                row_data[col_letter] = cell_value
        result.append(row_data)

    workbook.close()
    return result


import gspread
from oauth2client.service_account import ServiceAccountCredentials
import string

def column_letter_to_index(col_letter: str) -> int:
    """Преобразует буквы (например, A, B, AA) в индекс (0-based)"""
    col_letter = col_letter.upper()
    result = 0
    for c in col_letter:
        if c in string.ascii_uppercase:
            result = result * 26 + (ord(c) - ord('A') + 1)
        else:
            raise ValueError(f"Недопустимый символ в названии столбца: {c}")
    return result - 1  # Приводим к 0-based

def read_ppr_google_sheet_data(**kwargs) -> dict:
    """
    Считывает выбранные столбцы из Google Sheets.
    Аргументы через kwargs:
        - sheet_url: ссылка на таблицу
        - sheet_name: имя листа
        - creds_path: путь к JSON (по умолчанию 'credentials.json')
        - start_row: с какой строки читать (по умолчанию 2)
        - task_id, date, raw_location, problem, solution, blank: буквенные имена столбцов (например, "A", "C", "E")
    """

    # Настройки по умолчанию
    creds_path = kwargs.get("creds_path", "credentials.json")
    sheet_url = kwargs.get("sheet_url")
    sheet_name = kwargs.get("sheet_name")
    start_row = kwargs.get("start_row", 2)

    # Проверка обязательных параметров
    if not sheet_url or not sheet_name:
        return {"success": False, "message": "Отсутствует sheet_url или sheet_name"}

    # Обработка словаря столбцов
    columns_to_read = {
        "ppr_short_name": kwargs.get("ppr_short_name"),
        "ppr_address": kwargs.get("ppr_address"),
        "ppr_ID1": kwargs.get("ppr_ID1"),
        "ppr_ID2": kwargs.get("ppr_ID2")
    }

    if not all(columns_to_read.values()):
        return {
            "success": False,
            "message": "Некорректно указаны столбцы: все поля (task_id, date и т.д.) должны быть заданы"
        }
    # Конвертация букв в индексы
    try:
        col_indices = {
            key: column_letter_to_index(val)
            for key, val in columns_to_read.items()
        }
    except Exception as e:
        return {"success": False, "message": f"Ошибка при обработке буквенных названий столбцов: {e}"}

    # Авторизация
    try:
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
        client = gspread.authorize(creds)
    except Exception as e:
        return {"success": False, "message": f"Ошибка авторизации: {e}"}

    # Открытие листа
    try:
        sheet = client.open_by_url(sheet_url).worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound as e:
        return {"success": False, "message": f"Лист '{sheet_name}' не найден: {e}"}
    except Exception as e:
        return {"success": False, "message": f"Ошибка открытия таблицы: {e}"}

    # Получение всех данных
    try:
        all_data = sheet.get_all_values()
    except Exception as e:
        return {"success": False, "message": f"Ошибка получения данных с листа: {e}"}

    if len(all_data) < start_row:
        return {"success": False, "message": "Недостаточно строк на листе"}

    # Сбор нужных данных
    selected_data = []
    try:
        for row in all_data[start_row - 1:]:  # -1 т.к. индексация с 0
            if all(cell == "" for cell in row):
                break
            row_data = {
                key: row[col_indices[key]] if col_indices[key] < len(row) else ""
                for key in col_indices
            }
            selected_data.append(row_data)
    except Exception as e:
        return {"success": False, "message": f"Ошибка обработки данных: {e}"}

    return {"success": True, "data": selected_data}

def read_google_sheet_data(**kwargs) -> dict:
    """
    Считывает выбранные столбцы из Google Sheets.
    Аргументы через kwargs:
        - sheet_url: ссылка на таблицу
        - sheet_name: имя листа
        - creds_path: путь к JSON (по умолчанию 'credentials.json')
        - start_row: с какой строки читать (по умолчанию 2)
        - task_id, date, raw_location, problem, solution, blank: буквенные имена столбцов (например, "A", "C", "E")
    """

    # Настройки по умолчанию
    creds_path = kwargs.get("creds_path", "credentials.json")
    sheet_url = kwargs.get("sheet_url")
    sheet_name = kwargs.get("sheet_name")
    start_row = kwargs.get("start_row", 2)

    # Проверка обязательных параметров
    if not sheet_url or not sheet_name:
        return {"success": False, "message": "Отсутствует sheet_url или sheet_name"}

    # Обработка словаря столбцов
    columns_to_read = {
        "task_id": kwargs.get("task_id"),
        "date": kwargs.get("date"),
        "raw_location": kwargs.get("raw_location"),
        "problem": kwargs.get("problem"),
        "solution": kwargs.get("solution"),
        "blank": kwargs.get("blank")
    }

    if not all(columns_to_read.values()):
        return {
            "success": False,
            "message": "Некорректно указаны столбцы: все поля (task_id, date и т.д.) должны быть заданы"
        }
    # Конвертация букв в индексы
    try:
        col_indices = {
            key: column_letter_to_index(val)
            for key, val in columns_to_read.items()
        }
    except Exception as e:
        return {"success": False, "message": f"Ошибка при обработке буквенных названий столбцов: {e}"}

    # Авторизация
    try:
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
        client = gspread.authorize(creds)
    except Exception as e:
        return {"success": False, "message": f"Ошибка авторизации: {e}"}

    # Открытие листа
    try:
        sheet = client.open_by_url(sheet_url).worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound as e:
        return {"success": False, "message": f"Лист '{sheet_name}' не найден: {e}"}
    except Exception as e:
        return {"success": False, "message": f"Ошибка открытия таблицы: {e}"}

    # Получение всех данных
    try:
        all_data = sheet.get_all_values()
    except Exception as e:
        return {"success": False, "message": f"Ошибка получения данных с листа: {e}"}

    if len(all_data) < start_row:
        return {"success": False, "message": "Недостаточно строк на листе"}

    # Сбор нужных данных
    selected_data = []
    try:
        for row in all_data[start_row - 1:]:  # -1 т.к. индексация с 0
            if all(cell == "" for cell in row):
                break
            row_data = {
                key: row[col_indices[key]] if col_indices[key] < len(row) else ""
                for key in col_indices
            }
            selected_data.append(row_data)
    except Exception as e:
        return {"success": False, "message": f"Ошибка обработки данных: {e}"}

    return {"success": True, "data": selected_data}

import random

def generate_ppr_id(home_id):
    new_id = None
    with Session() as dbsession:
        ids = [row[0] for row in dbsession.query(TaskModel.task_id).filter(
            TaskModel.home_id.in_([home_id]),
            TaskModel.is_ppr == 1
        ).all()]

        # Получаем все существующие task_id (в виде множества для быстрого поиска)
        existing_ids = set(
            dbsession.query(TaskModel.task_id).all()
        )
        existing_ids = {row[0] for row in existing_ids}  # преобразуем в set int
    if ids:
        print('exist')
        return ids[0]
    # Генерация уникального id
    while True:
        new_id = random.randint(1, 10**12)  # или свой диапазон
        if new_id not in existing_ids:
            return new_id

def main(**kwargs):
    incorrect_addresses = []
    submit_required = []
    already_existing = []
    if kwargs.get("from_exel", False):
        excel_file = kwargs.get('file', False)
        sheet_name = kwargs.get('sheet', False)
        if not (bool(excel_file) and bool(sheet_name)):
            return {"success": False,
                    "message": "не передано название файла или название листа",
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing
                    }
        columns_to_read = {"task_id": kwargs.get("task_id", False),
                           "date": kwargs.get("date", False),
                           "raw_location": kwargs.get("raw_location", False),
                           "problem": kwargs.get("problem", False),
                           "solution":kwargs.get("solution", False),
                           "blank": kwargs.get("blank", False)
        }
        if not all([columns_to_read[i] for i in columns_to_read]):
            return {"success": False,
                    "message": "указаны некорректные данные шапки таблицы",
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing
                    }
        start_row = kwargs.get("start_string", 2)

        try:
            head = ['task_id', 'date', 'raw_location', 'problem', 'solution', 'blank']
            data = read_selected_columns(file_path=excel_file,
                                         sheet_name=sheet_name,
                                         start_row=start_row,
                                         end_row=None,
                                         columns_to_read=[columns_to_read[i] for i in head])
        except Exception as e:
            return {"success": False,
                    "message": f"невозможно считать выбранный лист из файла. Ошибка: {str(e.args)}",
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing
                    }
        task_id_header = kwargs.get("task_id", False)
        for i, row in enumerate(data, start=1):

            if all(str(cell).strip() == '' for cell in row.values()):
                break
            if not row[task_id_header]:
                return {
                    "success": False,
                 "message": "нет возможности считать ID задачи",
                        "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing}
            address = Address(
                raw_address=row[columns_to_read["raw_location"]]
            )

            if address.is_valid and not address.submit_required:
                task = Task(task_id=int(row[columns_to_read["task_id"]]),
                     date=row[columns_to_read["date"]],
                     address=address,
                     problem=row[columns_to_read["problem"]],
                     solution=row[columns_to_read["solution"]],
                     blank=bool(row[columns_to_read["blank"]] == kwargs.get("checked_value", "")),
                    ppr=False
                     )
                if task.already_exist and task.blank:
                    already_existing.append({
                        "task_id": task.task_id,
                        "blank": task.blank,
                        "archieved": task.archieved
                    })
            elif not address.is_valid:
                incorrect_addresses.append(row[columns_to_read["raw_location"]])
            if address.submit_required and address.is_valid:
                submit_required.append(address.__dict__())
        if not incorrect_addresses and not submit_required:
            return {"success": True,
                    "message": "задачи загружены",
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing}

        return {"success": False,
                "message": "есть нераспознанные адреса или адреса, требующие подтверждения",
                "incorrect_addresses": incorrect_addresses,
                "submit_required": submit_required,
                "already_existing": already_existing
                }
    elif kwargs.get("from_link", False):
        columns_to_read = {"task_id": kwargs.get("task_id", False),
                           "date": kwargs.get("date", False),
                           "raw_location": kwargs.get("raw_location", False),
                           "problem": kwargs.get("problem", False),
                           "solution": kwargs.get("solution", False),
                           "blank": kwargs.get("blank", False)
                           }
        if not all([columns_to_read[i] for i in columns_to_read]):
            return {"success": False,
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing,
                    "message": "указаны некорректные данные шапки таблицы"
                    }
        result = read_google_sheet_data(
            sheet_url=kwargs.get("from_link", ""),
            sheet_name=kwargs.get('sheet', False),
            task_id=kwargs.get('task_id', False),
            date=kwargs.get('date', False),
            raw_location=kwargs.get('raw_location', False),
            problem=kwargs.get('problem', False),
            solution=kwargs.get('solution', False),
            blank=kwargs.get('blank', False),
            start_row=kwargs.get('start_row', 2),
            creds_path=creds_path
        )
        if result["success"]:

            for i, row in enumerate(result['data'], start=1):
                if all(str(cell).strip() == '' for cell in row.values()):
                    break
                if not row["task_id"]:
                    return {"success": False,
                            "message": "нет возможности считать ID задачи",
                            "incorrect_addresses": incorrect_addresses,
                            "submit_required": submit_required,
                            "already_existing": already_existing
                            }
                address = Address(
                    raw_address=row["raw_location"]
                )

                if address.is_valid and not address.submit_required:
                    task = Task(task_id=int(row["task_id"]),
                         date=row["date"],
                         address=address,
                         problem=row["problem"],
                         solution=row["solution"],
                         blank=bool(row["blank"] == kwargs.get("checked_value", "Да")),
                         ppr=False
                         )

                    if task.already_exist and  task.blank:
                        already_existing.append({
                            "task_id": task.task_id,
                            "blank": task.blank,
                            "archieved": task.archieved
                        })
                elif not address.is_valid:
                    incorrect_addresses.append(row["raw_location"])
                if address.submit_required and address.is_valid:
                    submit_required.append(address.__dict__())
            if not incorrect_addresses and not submit_required:
                return {"success": True,
                        "message": "задачи загружены",
                        "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                        "already_existing": already_existing}

            return {"success": False,
                    "message": "есть нераспознанные адреса или адреса, требующие подтверждения",
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing}
        else:
            return {"success": False,
                    "message": result["message"],
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing
                    }

    return {"success": False,
            "message": "указан невалидный способ получения данных",
            "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
            "already_existing": already_existing}


def make_ppr(**kwargs):
    incorrect_addresses = []
    submit_required = []
    already_existing = []
    result = read_ppr_google_sheet_data(
        sheet_url=kwargs.get("from_link", ""),
        sheet_name=kwargs.get('sheet', False),
        ppr_short_name=kwargs.get('ppr_short_name', False),
        ppr_address=kwargs.get('ppr_address', False),
        ppr_ID1=kwargs.get('ppr_ID1', False),
        ppr_ID2=kwargs.get('ppr_ID2', False),
        creds_path=creds_path
    )
    if result["success"]:
        for i, row in enumerate(result['data'], start=1):
            if all(str(cell).strip() == '' for cell in row.values()):
                break
            """if not row["task_id"]:
                return {"success": False,
                        "message": "нет возможности считать ID задачи",
                        "incorrect_addresses": incorrect_addresses,
                        "submit_required": submit_required,
                        "already_existing": already_existing
                        }"""
            address = Address(
                raw_address=row["ppr_ID1"] + " " + row["ppr_short_name"] + " " + row["ppr_address"]
            )

            if address.is_valid and not address.submit_required:
                task = Task(task_id=generate_ppr_id(address.home_id),
                            date="",
                            address=address,
                            problem=(
                                    f"ППР"
                                    + '\n<br>'
                                    + f"ИД1: {row['ppr_ID1']}"
                                    + (int(bool(row['ppr_ID2'])) * ('\n<br>' + f"ИД2: {row['ppr_ID2']}"))

                            ),

                solution='',
                            blank=False,
                            ppr=True
                            )

                if task.already_exist and task.blank:
                    already_existing.append({
                        "task_id": task.task_id,
                        "blank": task.blank,
                        "archieved": task.archieved
                    })
            elif not address.is_valid:
                incorrect_addresses.append(row["raw_location"])
            if address.submit_required and address.is_valid:
                submit_required.append(address.__dict__())
        if not incorrect_addresses and not submit_required:
            return {"success": True,
                    "message": "задачи загружены",
                    "incorrect_addresses": incorrect_addresses,
                    "submit_required": submit_required,
                    "already_existing": already_existing}

        return {"success": False,
                "message": "есть нераспознанные адреса или адреса, требующие подтверждения",
                "incorrect_addresses": incorrect_addresses,
                "submit_required": submit_required,
                "already_existing": already_existing}
    else:
        return {"success": False,
                "message": result["message"],
                "incorrect_addresses": incorrect_addresses,
                "submit_required": submit_required,
                "already_existing": already_existing
                }


if __name__ == "__main__":
    print(main(from_link="https://docs.google.com/spreadsheets/d/1i6XreBzCqaFlfUAAJ2DRBZO2HrxnPDzh3HDIpl6fWH8/",
               ))
    #print(Address(home_id=4029).is_exist(4029))
    #print(Address(raw_address='2694 ГБОУ Школа "Марьино" Донецкая улица, дом 6'))
    #print(Address(home_id=723))
    #print(Address(raw_address="ZOV"))
    #print(Address(home_id="zov"))
    #main(User("NN", "123"), from_exel=False)
    #main(User("NN2", "123"), from_exel=False)
